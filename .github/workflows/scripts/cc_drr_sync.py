#!/usr/bin/env python3
"""
CC DRR Sync Script
Fetches the CC DRR Google Sheet as XLSX, flattens all platform tabs,
and upserts to Supabase drr_data table.

Runs via GitHub Actions nightly or on manual trigger.
"""

import os
import io
import json
import time
import requests
import openpyxl
from datetime import datetime, date

# ─────────────────────────────────────────────
# CONFIG — all secrets come from environment variables
# set in GitHub repo Settings → Secrets
# ─────────────────────────────────────────────
SHEET_ID        = '1Gbu3bSuhYJjo3oUf4ZU8GdA0YPDsFafBtxYECN9L6n8'  # CC DRR mirror sheet
SUPABASE_URL    = os.environ['SUPABASE_URL']
SUPABASE_KEY    = os.environ['SUPABASE_KEY']
SUPABASE_TABLE  = 'drr_data'
DATE_FROM       = '2025-11-01'  # only sync data from this date onwards

PLATFORM_CONFIGS = {
    'Amazon':   {'sku_col': 0, 'data_start': 3},
    'Flipkart': {'sku_col': 0, 'data_start': 4},
    'Myntra':   {'sku_col': 0, 'data_start': 4},
    'Nykaa':    {'sku_col': 0, 'data_start': 3},
    'Swiggy':   {'sku_col': 0, 'data_start': 3},
    'Blinkit':  {'sku_col': 0, 'data_start': 4},
    'Meesho':   {'sku_col': 0, 'data_start': 2},
    'Website':  {'sku_col': 0, 'data_start': 4},
    'Zepto':    {'sku_col': 0, 'data_start': 3, 'aggregate': True},
}

HEADERS = {
    'apikey':        SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type':  'application/json',
    'Prefer':        'resolution=merge-duplicates',
}

# ─────────────────────────────────────────────
# STEP 1 — Download XLSX from Google Sheets
# ─────────────────────────────────────────────
def download_sheet():
    url = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'
    print(f'Downloading sheet: {url}')
    res = requests.get(url, timeout=120)
    if res.status_code != 200:
        raise Exception(f'Sheet download failed: HTTP {res.status_code}')
    print(f'Downloaded {len(res.content) / 1024:.0f} KB')
    return openpyxl.load_workbook(io.BytesIO(res.content), data_only=True)

# ─────────────────────────────────────────────
# STEP 2 — Load SKU Master
# ─────────────────────────────────────────────
def load_sku_master(wb):
    ws  = wb['SKU Master']
    map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0] or '').strip()
        name = str(row[1] or '').strip()
        if code and code.startswith('CC'):
            map[code] = name or code
    print(f'SKU Master: {len(map)} entries')
    return map

# ─────────────────────────────────────────────
# STEP 3 — Parse platform sheets
# ─────────────────────────────────────────────
def fmt_date(d):
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)
    return None

def parse_platform(wb, name, cfg):
    if name not in wb.sheetnames:
        print(f'WARNING: sheet "{name}" not found — skipping')
        return []

    ws     = wb[name]
    header = [c.value for c in ws[1]]

    # Build date column index — only >= DATE_FROM
    date_idx = {}
    for i, v in enumerate(header):
        if i < cfg['data_start']: continue
        d = fmt_date(v) if isinstance(v, (datetime, date)) else None
        if d and d >= DATE_FROM:
            date_idx[i] = d

    if not date_idx:
        print(f'{name}: no date columns found after {DATE_FROM}')
        return []

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[cfg['sku_col']] or '').strip()
        if not sku or not sku.startswith('CC'): continue
        for col, ds in date_idx.items():
            val = row[col]
            if val is None: continue
            try:
                units = float(val)
                if units > 0:
                    records.append({'platform': name, 'sku': sku, 'date': ds, 'units': round(units)})
            except (TypeError, ValueError):
                pass

    return records

def parse_zepto(wb, cfg):
    name = 'Zepto'
    if name not in wb.sheetnames:
        print(f'WARNING: Zepto sheet not found — skipping')
        return []

    ws     = wb[name]
    header = [c.value for c in ws[1]]

    date_idx = {}
    for i, v in enumerate(header):
        if i < cfg['data_start']: continue
        d = fmt_date(v) if isinstance(v, (datetime, date)) else None
        if d and d >= DATE_FROM:
            date_idx[i] = d

    agg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[cfg['sku_col']] or '').strip()
        if not sku or not sku.startswith('CC'): continue
        for col, ds in date_idx.items():
            val = row[col]
            if val is None: continue
            try:
                units = float(val)
                if units > 0:
                    key      = f'{sku}||{ds}'
                    agg[key] = agg.get(key, 0) + units
            except (TypeError, ValueError):
                pass

    return [
        {'platform': 'Zepto', 'sku': k.split('||')[0], 'date': k.split('||')[1], 'units': round(v)}
        for k, v in agg.items()
    ]

# ─────────────────────────────────────────────
# STEP 4 — Upsert to Supabase in batches
# ─────────────────────────────────────────────
def upsert_to_supabase(records, sku_master):
    endpoint  = f'{SUPABASE_URL}/rest/v1/{SUPABASE_TABLE}'
    BATCH     = 500
    total     = len(records)
    done      = 0
    errors    = 0

    print(f'Upserting {total:,} records to Supabase in batches of {BATCH}…')

    for i in range(0, total, BATCH):
        batch = [
            {
                'platform': r['platform'],
                'sku':      r['sku'],
                'sku_name': sku_master.get(r['sku'], r['sku']),
                'date':     r['date'],
                'units':    r['units'],
            }
            for r in records[i:i + BATCH]
        ]

        res = requests.post(endpoint, headers=HEADERS, json=batch, timeout=30)

        if res.status_code not in (200, 201):
            print(f'  Batch {i//BATCH + 1} ERROR: HTTP {res.status_code} — {res.text[:200]}')
            errors += 1
        else:
            done += len(batch)
            pct   = round(done / total * 100)
            print(f'  Batch {i//BATCH + 1}: {done:,}/{total:,} rows ({pct}%)')

        time.sleep(0.1)  # be gentle with the API

    return done, errors

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    start = time.time()
    print(f'\n{"="*60}')
    print(f'CC DRR Sync — {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} UTC')
    print(f'{"="*60}\n')

    # 1. Download
    wb = download_sheet()

    # 2. SKU Master
    sku_master = load_sku_master(wb)

    # 3. Parse all platforms
    all_records = []
    for name, cfg in PLATFORM_CONFIGS.items():
        if cfg.get('aggregate'):
            recs = parse_zepto(wb, cfg)
        else:
            recs = parse_platform(wb, name, cfg)
        print(f'{name}: {len(recs):,} records')
        all_records.extend(recs)

    print(f'\nTotal records: {len(all_records):,}')

    if not all_records:
        print('ERROR: No records parsed — aborting')
        exit(1)

    # 4. Upsert
    done, errors = upsert_to_supabase(all_records, sku_master)

    elapsed = time.time() - start
    print(f'\n{"="*60}')
    print(f'Done in {elapsed:.1f}s — {done:,} rows synced, {errors} errors')
    print(f'{"="*60}\n')

    if errors > 0:
        exit(1)

if __name__ == '__main__':
    main()
